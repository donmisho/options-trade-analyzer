"""
Seed importer: Scoring Parameters.xlsx → engine_* configuration tables.

STANDALONE, BUILD-TIME ONLY.
    The workbook is the one-time seed source. After import, the runtime tables
    are the source of truth and are maintained through the strategy-administration
    UI. Re-import is for dev rebuilds only — it will UPSERT (never duplicate)
    on natural keys. Production edits live in the tables, NOT in the spreadsheet.

Usage:
    cd <project-root>
    python -m scripts.seed_engine_config [--xlsx path/to/workbook.xlsx] [--dry-run]

Requires:
    - OTA-681 engine_* tables already applied (alembic upgrade head)
    - openpyxl installed (pip install openpyxl)
    - Azure SQL reachable via Entra ID (az login)
"""

import argparse
import json
import logging
import re
import struct
import sys
import urllib.parse
from pathlib import Path

import openpyxl
import pyodbc

# Add project root to path so we can import app.core.config
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.core.config import settings

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Default workbook path ────────────────────────────────────────────────
DEFAULT_XLSX = Path(__file__).resolve().parent.parent / "requirements" / "Configuration" / "Scoring Parameters.xlsx"

# ── Strategy column mapping (0-indexed column letters → indices) ─────────
# Each strategy occupies 4 columns: Include, Low, High, Weight
STRATEGIES = [
    {"key": "steady_paycheck",  "display": "Steady Paycheck",  "cols": {"include": 6, "low": 7, "high": 8, "weight": 9}},   # G H I J
    {"key": "weekly_grind",     "display": "Weekly Grind",      "cols": {"include": 10, "low": 11, "high": 12, "weight": 13}}, # K L M N
    {"key": "trend_rider",      "display": "Trend Rider",       "cols": {"include": 14, "low": 15, "high": 16, "weight": 17}}, # O P Q R
    {"key": "lottery_ticket",   "display": "Lottery Ticket",    "cols": {"include": 18, "low": 19, "high": 20, "weight": 21}}, # S T U V
]

# ── Scan parameters from rows 69-72 (per strategy, same column order) ───
SCAN_PARAM_ROWS = {
    69: "evaluation_cap",
    70: "dte_priority_start",
    71: "dte_expansion_direction",
    72: "output_rank_cap",
}
SCAN_PARAM_COLS = {
    "steady_paycheck": 6,   # G
    "weekly_grind": 10,      # K
    "trend_rider": 14,       # O
    "lottery_ticket": 18,    # S
}


def slugify(text: str) -> str:
    """Convert a description to a snake_case rule_key."""
    text = text.strip().lower()
    text = re.sub(r"[–—]", "_", text)       # em/en dashes
    text = re.sub(r"[^a-z0-9_]+", "_", text) # non-alnum → underscore
    text = re.sub(r"_+", "_", text)          # collapse runs
    return text.strip("_")


def validate_json(value, context: str) -> str:
    """Validate and return a JSON string. Raises on invalid JSON."""
    if value is None:
        return None
    s = json.dumps(value) if not isinstance(value, str) else value
    try:
        json.loads(s)
    except (json.JSONDecodeError, TypeError) as e:
        raise ValueError(f"Invalid JSON for {context}: {e}\n  Value: {s!r}")
    return s


def connect_db() -> pyodbc.Connection:
    """Connect to Azure SQL using the project's DATABASE_URL + Entra ID token."""
    database_url = settings.database_url
    if not database_url.startswith("mssql"):
        raise RuntimeError(f"Expected mssql DATABASE_URL, got: {database_url[:30]}...")

    parsed = urllib.parse.urlparse(database_url)
    server = parsed.hostname
    port = parsed.port or 1433
    database = parsed.path.lstrip("/")

    odbc_connect = (
        f"Driver={{ODBC Driver 18 for SQL Server}};"
        f"Server={server},{port};"
        f"Database={database};"
        f"Encrypt=yes;"
        f"TrustServerCertificate=no;"
    )

    from azure.identity import DefaultAzureCredential
    cred = DefaultAzureCredential()
    token = cred.get_token("https://database.windows.net/.default")
    token_bytes = token.token.encode("UTF-16-LE")
    token_struct = struct.pack(f"<I{len(token_bytes)}s", len(token_bytes), token_bytes)

    SQL_COPT_SS_ACCESS_TOKEN = 1256
    conn = pyodbc.connect(odbc_connect, attrs_before={SQL_COPT_SS_ACCESS_TOKEN: token_struct})
    conn.autocommit = False
    log.info(f"Connected to {server}/{database}")
    return conn


# ── Workbook parsing ─────────────────────────────────────────────────────

def parse_workbook(xlsx_path: Path):
    """Parse the workbook into rule dicts, strategy dicts, junction dicts, lookup dicts."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Sheet1"]

    # Read all rows into a list of dicts (0-indexed columns)
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        all_rows.append(list(row))

    rules = []
    junctions = []  # list of (rule_key, strategy_key, junction_data)
    eval_order_counters = {}  # (strategy_key, phase) → next order

    def next_eval_order(strategy_key, phase):
        k = (strategy_key, phase)
        eval_order_counters[k] = eval_order_counters.get(k, 0) + 1
        return eval_order_counters[k]

    # Parse data rows (4 through 58, skipping blanks and section headers)
    for row_idx in range(3, len(all_rows)):  # 0-indexed, so row 4 = index 3
        row = all_rows[row_idx]
        rule_type = row[0]  # col A
        if not rule_type or rule_type in (
            "Pattern: Rules below may also appear as Hard Gates above. This is intentional layering",
            "POST-SCORING ADJUSTMENTS",
            "VERDICT BANDS",
            "STRATEGY SCAN PARAMETERS",
        ):
            continue
        # Stop before verdict bands section
        if str(rule_type).startswith("Pattern:"):
            continue
        if rule_type in ("Verdict", "EXECUTE", "WAIT", "PASS", "Note:", "Parameter"):
            continue
        if rule_type not in ("Hard Gate", "Soft Gate", "Scoring Criteria", "Post-Scoring"):
            continue

        category = row[1]   # col B
        description = row[2]  # col C
        dtype = row[3]       # col D
        condition = row[4]   # col E
        tier = row[5]        # col F

        if not description:
            continue

        rule_key = slugify(str(description))

        # Map rule_type to engine phase
        if rule_type in ("Hard Gate", "Soft Gate"):
            phase = "gate"
        elif rule_type == "Scoring Criteria":
            phase = "scoring"
        elif rule_type == "Post-Scoring":
            phase = "adjustment"
            # Prefix adjustment rules to avoid key collision with gate rules
            # that share the same description (intentional layering per workbook)
            rule_key = f"adj_{rule_key}"
        else:
            continue

        # Determine null_semantics for data-completeness gates
        null_semantics = None
        if description and "Data completeness" in str(description):
            null_semantics = "FAIL_CLOSED"

        # Determine condition_expression
        condition_expression = str(condition) if condition else None

        # Determine formula_ref
        formula_ref = None
        if condition and "Black Scholes" in str(condition):
            formula_ref = f"formula:{rule_key}"
        elif condition and str(condition).startswith("TBD"):
            formula_ref = f"formula:{rule_key}"

        # Build referenced_named_values from condition text (best effort)
        ref_values = _extract_named_values(condition, description)

        # Build parameter_schema based on what parameters this rule accepts
        param_schema = _build_parameter_schema(rule_type, condition, description, dtype)

        rule = {
            "owner_app_id": "OTA",
            "rule_key": rule_key,
            "phase": phase,
            "tier": str(tier) if tier else None,
            "intent": str(description),
            "condition_expression": condition_expression,
            "formula_ref": formula_ref,
            "referenced_named_values": ref_values,
            "parameter_schema": param_schema,
            "null_semantics": null_semantics,
            "enabled": True,
        }
        rules.append(rule)

        # Build junction rows for each strategy
        for strat in STRATEGIES:
            cols = strat["cols"]
            include = row[cols["include"]]
            low_val = row[cols["low"]]
            high_val = row[cols["high"]]
            weight_val = row[cols["weight"]]

            # Scoring criteria: participation is indicated by a non-null weight,
            # not by the Include column. Gates/adjustments use Include = "Y".
            if phase == "scoring":
                if weight_val is None:
                    continue
            else:
                if not include or str(include).upper() != "Y":
                    continue

            # Build parameters JSON
            params = {}
            if low_val is not None and str(low_val) not in ("N/A", "lookup", ""):
                try:
                    params["low"] = float(low_val)
                except (ValueError, TypeError):
                    params["low"] = str(low_val)
            if high_val is not None and str(high_val) not in ("N/A", "lookup", ""):
                try:
                    params["high"] = float(high_val)
                except (ValueError, TypeError):
                    params["high"] = str(high_val)

            # For lookup-based thresholds (width tier compliance)
            if low_val == "lookup" or high_val == "lookup":
                params["lookup_set"] = "width_configuration"

            # Determine stop_if_fail and score_penalty
            if rule_type == "Hard Gate":
                stop_if_fail = True
                score_penalty = None
            elif rule_type == "Soft Gate":
                stop_if_fail = False
                score_penalty = _extract_penalty(condition)
            elif rule_type == "Post-Scoring":
                stop_if_fail = False
                score_penalty = _extract_penalty(condition)
            else:
                stop_if_fail = False
                score_penalty = None

            # Weight (scoring criteria only)
            weight = None
            if phase == "scoring" and weight_val is not None:
                try:
                    weight = float(weight_val)
                except (ValueError, TypeError):
                    pass

            junction = {
                "rule_key": rule_key,
                "strategy_key": strat["key"],
                "evaluation_order": next_eval_order(strat["key"], phase),
                "stop_if_fail": stop_if_fail,
                "score_penalty": score_penalty,
                "weight": weight,
                "parameters": params if params else None,
                "terminal_verdict": None,
                "enabled": True,
            }
            junctions.append(junction)

    # ── Post-process: enrich ETF adjustment rule's referenced_named_values ──
    # OTA-690: is_etf is a required input (producer = options-chain adapter,
    # not yet built). The ETF rule also references the width_configuration
    # lookup for ETF-specific spread width adjustments.
    for r in rules:
        if r["rule_key"].startswith("adj_") and "etf" in r["rule_key"]:
            r["referenced_named_values"] = [
                {
                    "name": "is_etf",
                    "producer": "options-chain-adapter",
                    "status": "requirement-only",
                    "note": "Producer not yet built — later feature. Records the input requirement.",
                },
                {
                    "name": "width_configuration",
                    "type": "lookup_ref",
                    "lookup_set": "width_configuration",
                    "note": "Width tiers may vary for ETF underlyings; runtime resolves via lookup.",
                },
            ]

    # ── Strategies ───────────────────────────────────────────────────────
    # Per-strategy verdict bands (OTA-773). Each strategy gets its own copy
    # so bands can diverge independently. All start at 70/50 (matching
    # _assign_verdict literals that OTA-761 will remove).
    _SCREENING_VERDICT_BANDS = {
        "steady_paycheck": [
            {"verdict": "EXECUTE", "min_score": 70, "max_score": 100},
            {"verdict": "WAIT",    "min_score": 50, "max_score": 69.99},
            {"verdict": "PASS",    "min_score": 0,  "max_score": 49.99},
        ],
        "weekly_grind": [
            {"verdict": "EXECUTE", "min_score": 70, "max_score": 100},
            {"verdict": "WAIT",    "min_score": 50, "max_score": 69.99},
            {"verdict": "PASS",    "min_score": 0,  "max_score": 49.99},
        ],
        "trend_rider": [
            {"verdict": "EXECUTE", "min_score": 70, "max_score": 100},
            {"verdict": "WAIT",    "min_score": 50, "max_score": 69.99},
            {"verdict": "PASS",    "min_score": 0,  "max_score": 49.99},
        ],
        "lottery_ticket": [
            {"verdict": "EXECUTE", "min_score": 70, "max_score": 100},
            {"verdict": "WAIT",    "min_score": 50, "max_score": 69.99},
            {"verdict": "PASS",    "min_score": 0,  "max_score": 49.99},
        ],
    }

    # Scan parameters (rows 69-72)
    scan_params_by_strategy = {}
    for row_num, param_name in SCAN_PARAM_ROWS.items():
        row = all_rows[row_num - 1]  # 0-indexed
        for strat_key, col_idx in SCAN_PARAM_COLS.items():
            val = row[col_idx]
            if strat_key not in scan_params_by_strategy:
                scan_params_by_strategy[strat_key] = {}
            if val is not None:
                scan_params_by_strategy[strat_key][param_name] = val

    strategies = []
    for strat in STRATEGIES:
        # Extract DTE window from the per-strategy DTE hard gate (row 15 = index 14)
        dte_row = all_rows[14]
        cols = strat["cols"]
        dte_low = dte_row[cols["low"]]
        dte_high = dte_row[cols["high"]]

        # Determine compatible_structures from trade type gates
        compat = []
        for row_idx in range(23, 32):  # rows 24-32 (0-indexed 23-31)
            row = all_rows[row_idx]
            if row[0] == "Hard Gate" and row[1] == "Trade Type":
                include = row[cols["include"]]
                if include and str(include).upper() == "Y":
                    desc = str(row[2])
                    # Only add specific trade types, not meta-rules
                    if desc in ("BULL_PUT_CREDIT", "BEAR_CALL_CREDIT", "BULL_CALL_DEBIT",
                                "BEAR_PUT_DEBIT", "LONG_CALL", "LONG_PUT", "IRON_CONDOR"):
                        compat.append(desc)

        strategy = {
            "owner_app_id": "OTA",
            "strategy_key": strat["key"],
            "display_name": strat["display"],
            "consumer_surface": "SCREENING",
            "description": None,
            "compatible_structures": compat if compat else None,
            "verdict_band_set": _SCREENING_VERDICT_BANDS[strat["key"]],
            "dte_min": int(dte_low) if dte_low else None,
            "dte_max": int(dte_high) if dte_high else None,
            "enabled": True,
        }
        strategies.append(strategy)

    # ── Lookups ──────────────────────────────────────────────────────────
    lookups = []

    # Pipeline phases (SHARED)
    for i, phase in enumerate(["gate", "scoring", "adjustment", "verdict"], 1):
        lookups.append({
            "owner_app_id": "SHARED",
            "lookup_set": "pipeline_phases",
            "lookup_key": phase,
            "payload": {"label": phase.title(), "description": f"Pipeline phase: {phase}"},
            "sort_order": i,
        })

    # Calculation tiers (SHARED)
    for i, tier in enumerate(["RAW", "DERIVED", "COMPUTED"], 1):
        lookups.append({
            "owner_app_id": "SHARED",
            "lookup_set": "calculation_tiers",
            "lookup_key": tier,
            "payload": {"label": tier, "description": f"Calculation tier: {tier}"},
            "sort_order": i,
        })

    # Null semantics (SHARED)
    for i, sem in enumerate(["FAIL_OPEN", "FAIL_CLOSED", "SKIP"], 1):
        lookups.append({
            "owner_app_id": "SHARED",
            "lookup_set": "null_semantics",
            "lookup_key": sem,
            "payload": {"label": sem},
            "sort_order": i,
        })

    # Consumer surfaces (SHARED)
    for i, surf in enumerate(["SCREENING", "POSITION_HEALTH", "DIRECTIONAL"], 1):
        lookups.append({
            "owner_app_id": "SHARED",
            "lookup_set": "consumer_surfaces",
            "lookup_key": surf,
            "payload": {"label": surf},
            "sort_order": i,
        })

    # Verdict domains (OTA — screening verdicts: band verdicts)
    for i, v in enumerate(verdict_bands, 1):
        lookups.append({
            "owner_app_id": "OTA",
            "lookup_set": "screening_verdicts",
            "lookup_key": v["verdict"],
            "payload": {"min_score": v["min_score"], "max_score": v["max_score"]},
            "sort_order": i,
        })

    # OTA-711: Halt verdicts (bypass band lookup; emitted by terminal_verdict on junction)
    halt_verdicts = [
        {
            "lookup_key": "WAIT_FOR_EARNINGS",
            "payload": {
                "label": "Wait for Earnings",
                "kind": "HALT_VERDICT",
                "description": "Halt emitted by the earnings gate when an earnings "
                               "announcement falls inside the trade window.",
            },
        },
    ]
    for i, hv in enumerate(halt_verdicts, len(verdict_bands) + 1):
        lookups.append({
            "owner_app_id": "OTA",
            "lookup_set": "screening_verdicts",
            "lookup_key": hv["lookup_key"],
            "payload": hv["payload"],
            "sort_order": i,
        })

    # Width tiers from Width Configuration sheet (OTA)
    ws2 = wb["Width Configuration"]
    tier_count = 0
    for i, row in enumerate(ws2.iter_rows(min_row=3, max_row=7, values_only=True), 1):
        if row[0] is None:
            continue
        lookups.append({
            "owner_app_id": "OTA",
            "lookup_set": "width_configuration",
            "lookup_key": f"tier_{i}",
            "payload": {
                "price_min": row[0],
                "price_max": row[1],
                "width_min": row[2],
                "width_max": row[3],
                "width_increment": row[4],
            },
            "sort_order": i,
        })
        tier_count = i

    # Per-security override shape (empty — overrides added later via admin UI)
    lookups.append({
        "owner_app_id": "OTA",
        "lookup_set": "width_configuration",
        "lookup_key": "_override_schema",
        "payload": {
            "type": "per_security_override",
            "fields": {
                "ticker": {"type": "string", "required": True},
                "width_min": {"type": "number", "required": True},
                "width_max": {"type": "number", "required": True},
                "width_increment": {"type": "number", "required": True},
                "note": {"type": "string", "required": False},
            },
            "description": "Per-security width overrides. Ticker match → use override; else fall back to price-tier default.",
        },
        "sort_order": tier_count + 1,
    })

    # Scan parameters per strategy (OTA)
    for strat_key, params in scan_params_by_strategy.items():
        lookups.append({
            "owner_app_id": "OTA",
            "lookup_set": "scan_parameters",
            "lookup_key": strat_key,
            "payload": params,
            "sort_order": None,
        })

    return rules, strategies, junctions, lookups


def _extract_named_values(condition, description) -> list[str]:
    """Best-effort extraction of named values from condition text."""
    vals = []
    if not condition:
        return vals
    cond = str(condition)
    # Common patterns
    patterns = {
        r"\bstock_price\b": "stock_price",
        r"\bnext_earnings_date\b": "next_earnings_date",
        r"\bexpiry_date\b": "expiry_date",
        r"\bspread\b": "bid_ask_spread",
        r"\bmin_leg_OI\b": "min_leg_open_interest",
        r"\bmin_leg_volume\b": "min_leg_volume",
        r"\bDTE\b": "dte",
        r"\bIV_rank\b": "iv_rank",
        r"\bdelta\b": "delta",
        r"\bATR_14\b": "atr_14",
        r"\bcredit/spread_width\b": "credit_width_pct",
        r"\bdebit/spread_width\b": "debit_width_pct",
        r"\btotal_EV\b": "total_ev",
        r"\bspot\b": "stock_price",
        r"\bshort_strike\b": "short_strike",
        r"\bSMA_50\b": "sma_50",
        r"\bchart_state\b": "chart_state",
        r"\bspread_width\b": "spread_width",
        r"\bcushion_vs_ATR\b": "cushion_vs_atr",
        r"\btheta_total_over_trade\b": "theta_total",
        r"\bdebit\b": "debit",
    }
    for pattern, name in patterns.items():
        if re.search(pattern, cond, re.IGNORECASE):
            if name not in vals:
                vals.append(name)

    # Data completeness gates
    desc = str(description) if description else ""
    if "Data completeness" in desc:
        if "IV Rank" in desc:
            vals.append("iv_rank")
        elif "Delta" in desc:
            vals.append("delta")
        elif "ATR_14" in desc:
            vals.append("atr_14")

    # For ETF bonus
    if "ETF" in str(condition or ""):
        vals.append("is_etf")

    return vals


def _build_parameter_schema(rule_type, condition, description, dtype) -> dict | None:
    """Build a parameter schema for the rule based on its type."""
    if dtype == "Binary Choice":
        return None  # No parameters — it's a Y/N gate

    cond = str(condition) if condition else ""
    desc = str(description) if description else ""

    # Data completeness gates have no parameters
    if "IS NOT NULL" in cond or "IS NULL" in cond:
        return None

    # BETWEEN-style rules have low/high
    if "BETWEEN" in cond or "is BETWEEN" in cond:
        return {"low": {"type": "number"}, "high": {"type": "number"}}

    # >= threshold style
    if ">= threshold" in cond or "<= threshold" in cond:
        return {"low": {"type": "number"}, "high": {"type": "number"}}

    # Post-scoring penalties/bonuses with a threshold
    if rule_type == "Post-Scoring" and (">" in cond or "BETWEEN" in cond):
        return {"threshold": {"type": "number"}}

    # Scoring criteria — formulas, typically no user-configurable params
    if rule_type == "Scoring Criteria":
        return None

    # Chart state matching
    if "chart_state" in cond:
        return None

    # Lookup-based
    if "Width Configuration" in cond:
        return {"lookup_set": {"type": "string"}}

    return None


def _extract_penalty(condition) -> float | None:
    """Extract penalty/bonus amount from condition text like 'REDUCE SCORE BY 10'."""
    if not condition:
        return None
    m = re.search(r"(?:REDUCE|DECREASE)\s+SCORE\s+BY\s+(\d+)", str(condition), re.IGNORECASE)
    if m:
        return -float(m.group(1))
    m = re.search(r"INCREASE\s+SCORE\s+BY\s+(\d+)", str(condition), re.IGNORECASE)
    if m:
        return float(m.group(1))
    return None


# ── Compound rule decomposition (OTA-683) ───────────────────────────────
#
# Post-parse step: identifies compound rules in the workbook-parsed output
# and replaces each with its stated number of atomic rules (2/2/2/4/2/2 = 14).
# BETWEEN → two comparison rows. No runtime BETWEEN operator is implemented.
# Earnings 4-route tree mirrors earnings_gate.py route semantics exactly.
# Graduated cushion penalty mirrors strategy_scorer.py _cushion_penalty bands.


def _find_compound(rules, *keywords, phase=None, exclude=None):
    """Find a rule whose rule_key or intent contains all given keywords."""
    for r in rules:
        key = r["rule_key"].lower()
        intent = (r.get("intent") or "").lower()
        text = f"{key} {intent}"
        if all(kw.lower() in text for kw in keywords):
            if phase and r["phase"] != phase:
                continue
            if exclude and exclude.lower() in key:
                continue
            return r
    return None


def _make_atomic(base_rule, *, rule_key, intent, condition_expr=None,
                 formula_ref=None, ref_values=None, param_schema=None):
    """Create an atomic rule dict derived from a compound rule's metadata."""
    return {
        "owner_app_id": base_rule["owner_app_id"],
        "rule_key": rule_key,
        "phase": base_rule["phase"],
        "tier": base_rule["tier"],
        "intent": intent,
        "condition_expression": condition_expr,
        "formula_ref": formula_ref,
        "referenced_named_values": ref_values if ref_values is not None else base_rule.get("referenced_named_values", []),
        "parameter_schema": param_schema,
        "null_semantics": base_rule.get("null_semantics"),
        "enabled": True,
    }


# ── Individual decomposers ──────────────────────────────────────────────
# Each returns (old_rule_key, [(atomic_rule_dict, junction_overrides)]) or None.
# junction_overrides is a dict of fields to override on each replicated junction,
# or None to inherit the compound's junction properties unchanged.


def _decompose_chart_state(rules):
    """1. Chart state confirms direction → 2 atomic rules."""
    compound = _find_compound(rules, "chart_state", "direction", phase="gate")
    if not compound:
        compound = _find_compound(rules, "chart_state", "confirms", phase="gate")
    if not compound:
        return None

    atoms = [
        (_make_atomic(
            compound,
            rule_key="chart_state_valid_alignment",
            intent="Chart state must be Bullish Alignment or Bearish Alignment",
            condition_expr="chart_state IN ('Bullish Alignment', 'Bearish Alignment')",
            ref_values=["chart_state"],
        ), None),
        (_make_atomic(
            compound,
            rule_key="chart_state_matches_trade_direction",
            intent="Chart state alignment must match trade direction (bullish for bull trades, bearish for bear trades)",
            formula_ref="formula:chart_state_matches_direction",
            ref_values=["chart_state", "trade_direction"],
        ), None),
    ]
    return compound["rule_key"], atoms


def _decompose_stock_extended(rules):
    """2. Stock extended in trade direction → 2 atomic rules."""
    compound = _find_compound(rules, "stock", "extended", phase="adjustment")
    if not compound:
        return None

    atoms = [
        (_make_atomic(
            compound,
            rule_key="adj_stock_extended_magnitude",
            intent="Stock price extended > 5% from 50-day SMA",
            condition_expr="abs(spot - SMA_50) / SMA_50 > 0.05",
            ref_values=["stock_price", "sma_50"],
        ), None),
        (_make_atomic(
            compound,
            rule_key="adj_stock_extended_direction_match",
            intent="Extension direction matches trade direction (above SMA for bull, below for bear)",
            formula_ref="formula:extension_matches_trade_direction",
            ref_values=["stock_price", "sma_50", "trade_direction"],
        ), None),
    ]
    return compound["rule_key"], atoms


def _decompose_cushion_atr(rules):
    """3. Cushion barely above ATR floor (BETWEEN) → 2 comparison rules."""
    compound = _find_compound(rules, "cushion", "atr", phase="adjustment")
    if not compound:
        return None

    atoms = [
        (_make_atomic(
            compound,
            rule_key="adj_cushion_vs_atr_gte_floor",
            intent="Cushion-to-ATR ratio at or above lower bound (>= 1.0)",
            condition_expr="cushion_vs_ATR >= 1.0",
            ref_values=["cushion_vs_atr"],
            param_schema={"threshold": {"type": "number", "default": 1.0}},
        ), None),
        (_make_atomic(
            compound,
            rule_key="adj_cushion_vs_atr_lte_ceiling",
            intent="Cushion-to-ATR ratio at or below upper bound (<= 1.5)",
            condition_expr="cushion_vs_ATR <= 1.5",
            ref_values=["cushion_vs_atr"],
            param_schema={"threshold": {"type": "number", "default": 1.5}},
        ), None),
    ]
    return compound["rule_key"], atoms


def _decompose_earnings(rules):
    """4. Earnings gate 4-route tree → 4 atomic rules per earnings_gate.py.

    Route semantics mirror EarningsInWindowGate exactly:
      Route 1: dte_before <= 7, dte_after < 14  → PASS (stop)
      Route 2: dte_before <= 7, dte_after >= 14  → WAIT_FOR_EARNINGS (stop)
      Route 3: dte_before >= 8, dte_after >= 21  → WAIT_FOR_EARNINGS (stop)
      Route 4: dte_before >= 8, dte_after < 21   → score with 15-pt penalty (non-stop)
    """
    compound = _find_compound(rules, "earnings", phase="gate")
    if not compound:
        return None

    atoms = [
        (_make_atomic(
            compound,
            rule_key="earnings_route1_no_viable_window",
            intent="Route 1: No viable window — dte_before <= 7 and dte_after < 14. Verdict: PASS.",
            formula_ref="formula:earnings_route1_no_viable_window",
            ref_values=["next_earnings_date", "entry_date", "expiry_date",
                        "dte_before_earnings", "dte_after_earnings"],
        ), {"stop_if_fail": True, "score_penalty": None, "terminal_verdict": "PASS"}),

        (_make_atomic(
            compound,
            rule_key="earnings_route2_wait_post_window",
            intent="Route 2: Pre-earnings window too short, strong post-earnings window — dte_before <= 7 and dte_after >= 14. Verdict: WAIT_FOR_EARNINGS.",
            formula_ref="formula:earnings_route2_wait_post_window",
            ref_values=["next_earnings_date", "entry_date", "expiry_date",
                        "dte_before_earnings", "dte_after_earnings"],
        ), {"stop_if_fail": True, "score_penalty": None, "terminal_verdict": "WAIT_FOR_EARNINGS"}),

        (_make_atomic(
            compound,
            rule_key="earnings_route3_post_entry_better",
            intent="Route 3: Post-earnings entry likely better — dte_before >= 8 and dte_after >= 21. Verdict: WAIT_FOR_EARNINGS.",
            formula_ref="formula:earnings_route3_post_entry_better",
            ref_values=["next_earnings_date", "entry_date", "expiry_date",
                        "dte_before_earnings", "dte_after_earnings"],
        ), {"stop_if_fail": True, "score_penalty": None, "terminal_verdict": "WAIT_FOR_EARNINGS"}),

        (_make_atomic(
            compound,
            rule_key="earnings_route4_pre_momentum_play",
            intent="Route 4: Pre-earnings momentum play — dte_before >= 8 and dte_after < 21. Score normally with 15-point penalty, effective DTE = dte_before - 1.",
            formula_ref="formula:earnings_route4_pre_momentum_play",
            ref_values=["next_earnings_date", "entry_date", "expiry_date",
                        "dte_before_earnings", "dte_after_earnings"],
        ), {"stop_if_fail": False, "score_penalty": -15.0}),
    ]
    return compound["rule_key"], atoms


def _decompose_credit_debit(rules):
    """5. Credit/debit quality gate → 2 atomic rules.

    Thresholds from evaluation_routes.py lines 626-663:
      Credit: credit_pct_of_width >= 0.30 (fail below)
      Debit:  debit_pct_of_width  <= 0.40 (fail above)
    """
    compound = _find_compound(rules, "credit", "debit", phase="gate")
    if not compound:
        compound = _find_compound(rules, "credit", "width", phase="gate")
    if not compound:
        compound = _find_compound(rules, "spread", "quality", phase="gate")
    if not compound:
        return None

    atoms = [
        (_make_atomic(
            compound,
            rule_key="credit_pct_of_width_floor",
            intent="Credit spread: credit received must be >= 30% of spread width",
            condition_expr="credit_pct_of_width >= 0.30",
            ref_values=["credit_width_pct", "spread_width", "net_credit"],
            param_schema={"threshold": {"type": "number", "default": 0.30}},
        ), {"stop_if_fail": True, "score_penalty": None}),
        (_make_atomic(
            compound,
            rule_key="debit_pct_of_width_ceiling",
            intent="Debit spread: debit paid must be <= 40% of spread width",
            condition_expr="debit_pct_of_width <= 0.40",
            ref_values=["debit_width_pct", "spread_width", "net_debit"],
            param_schema={"threshold": {"type": "number", "default": 0.40}},
        ), {"stop_if_fail": True, "score_penalty": None}),
    ]
    return compound["rule_key"], atoms


def _decompose_cushion_penalty(rules):
    """6. Cushion penalty graduated bands → 2 ordered adjustment rules.

    Bands from strategy_scorer.py _cushion_penalty (lines 138-142):
      Band 1: cushion_pct < 1.0%  → -20 points
      Band 2: cushion_pct in [1.0%, 2.0%) → -10 points

    This compound exists in code (strategy_scorer.py) but may not be in the
    workbook as a single row. If no compound is found, the atomic rules are
    injected directly using a synthetic base template.
    """
    compound = _find_compound(rules, "cushion", phase="adjustment", exclude="atr")
    if not compound:
        compound = _find_compound(rules, "cushion", "penalty", phase="adjustment")

    # Synthetic base when the workbook has no matching compound row
    base = compound or {
        "owner_app_id": "OTA",
        "phase": "adjustment",
        "tier": "DERIVED",
        "null_semantics": None,
    }

    atoms = [
        (_make_atomic(
            base,
            rule_key="adj_cushion_penalty_severe",
            intent="Cushion below severe threshold — severe proximity to short strike (-20 points)",
            condition_expr="<",
            ref_values=["cushion_pct"],
            param_schema={"threshold": {"type": "number", "min": 0, "max": 100}},
        ), {"score_penalty": -20.0}),
        (_make_atomic(
            base,
            rule_key="adj_cushion_penalty_moderate",
            intent="Cushion in moderate band — moderate proximity to short strike (-10 points)",
            formula_ref="formula:cushion_penalty_moderate",
            ref_values=["cushion_pct"],
            param_schema={
                "lower_threshold": {"type": "number", "min": 0, "max": 100},
                "upper_threshold": {"type": "number", "min": 0, "max": 100},
            },
        ), {"score_penalty": -10.0}),
    ]

    if compound:
        return compound["rule_key"], atoms
    # No compound to remove — return sentinel key that won't match any existing rule
    return "_synthetic_cushion_penalty_", atoms


# ── Orchestrator ────────────────────────────────────────────────────────

def decompose_compound_rules(rules, junctions):
    """
    OTA-683: Decompose compound rules into atomic engine_rules rows.

    Six compounds are decomposed into 14 total atomic rules (2+2+2+4+2+2).
    BETWEEN-style conditions become two comparison rows; the runtime BETWEEN
    operator is deferred to the engine-core expression library.
    Earnings routes mirror earnings_gate.py; stop_if_fail is intrinsic per route.
    Graduated cushion penalty mirrors strategy_scorer.py _cushion_penalty bands.
    """
    decomposers = [
        ("chart_state_confirms_direction",    _decompose_chart_state),
        ("stock_extended_in_trade_direction",  _decompose_stock_extended),
        ("cushion_barely_above_atr_floor",     _decompose_cushion_atr),
        ("earnings_4_route_tree",              _decompose_earnings),
        ("credit_debit_quality_gate",          _decompose_credit_debit),
        ("cushion_penalty_graduated_bands",    _decompose_cushion_penalty),
    ]

    removals = set()
    additions = []       # (old_key, [(rule, junction_overrides)])
    injections = []      # [(rule, junction_overrides)] — new rules with no compound to replace

    for label, decomposer in decomposers:
        result = decomposer(rules)
        if result is None:
            log.warning(f"  Compound '{label}' not found in parsed rules — skipping")
            continue
        old_key, atoms = result
        atom_keys = [a[0]["rule_key"] for a in atoms]
        if old_key.startswith("_synthetic_"):
            # No compound to remove; rules are injected fresh
            injections.extend(atoms)
            log.info(f"  Injected (no workbook compound): {atom_keys}")
        else:
            removals.add(old_key)
            additions.append((old_key, atoms))
            log.info(f"  Decomposed '{old_key}' → {atom_keys}")

    # Additional compounds whose atoms already exist (produced by a sibling decomposer)
    # but that the decomposer doesn't catch directly. Remove from rules + junctions.
    extra_retirements = {"debit_of_width"}
    removals |= extra_retirements

    # Rules to keep disabled (unbind junctions but preserve the rule row)
    keep_disabled = {"days_until_next_earnings"}

    # Rebuild rules: remove compounds (except keep_disabled), add atomics + injections
    new_rules = []
    for r in rules:
        if r["rule_key"] in removals:
            if r["rule_key"] in keep_disabled:
                # Keep rule row but mark disabled — divergence record
                r["enabled"] = False
                new_rules.append(r)
            # else: fully removed (atoms replace it)
        else:
            new_rules.append(r)
    for _, atoms in additions:
        for rule, _ in atoms:
            new_rules.append(rule)
    for rule, _ in injections:
        new_rules.append(rule)

    # Rebuild junctions: replace compound refs with per-atomic refs
    new_junctions = []
    for j in junctions:
        if j["rule_key"] not in removals:
            new_junctions.append(j)
            continue
        for old_key, atoms in additions:
            if j["rule_key"] == old_key:
                for rule, overrides in atoms:
                    new_j = dict(j)
                    new_j["rule_key"] = rule["rule_key"]
                    if overrides:
                        new_j.update(overrides)
                    new_junctions.append(new_j)
                break

    # Create junctions for injected rules (no compound junction to expand).
    # Cushion penalty applies to credit-spread strategies (short strike present).
    if injections:
        _inject_junctions(new_junctions, injections)

    total_atomic = sum(len(a) for _, a in additions) + len(injections)
    log.info(f"Decomposition complete: {len(removals)} compounds → "
             f"{total_atomic} atomic rules ({len(injections)} injected fresh)")
    return new_rules, new_junctions, removals


# Strategies that use cushion-based adjustments (credit spreads with a short strike)
_CUSHION_PENALTY_STRATEGIES = ["steady_paycheck", "weekly_grind"]


# Per-strategy cushion penalty thresholds (OTA-770).
# SP and WG start with identical values; junction independence lets them diverge later.
_CUSHION_PENALTY_PARAMS = {
    "adj_cushion_penalty_severe": {
        "steady_paycheck": {"threshold": 1.0},
        "weekly_grind":    {"threshold": 1.0},
    },
    "adj_cushion_penalty_moderate": {
        "steady_paycheck": {"lower_threshold": 1.0, "upper_threshold": 2.0},
        "weekly_grind":    {"lower_threshold": 1.0, "upper_threshold": 2.0},
    },
}


def _inject_junctions(junctions, injections):
    """Create junction rows for freshly injected rules (no compound to expand from)."""
    for rule, overrides in injections:
        key = rule["rule_key"]
        if not key.startswith("adj_cushion_penalty_"):
            continue
        for strat_key in _CUSHION_PENALTY_STRATEGIES:
            params = _CUSHION_PENALTY_PARAMS.get(key, {}).get(strat_key)
            j = {
                "rule_key": key,
                "strategy_key": strat_key,
                "evaluation_order": 0,   # placeholder — OTA-684 sets proper ordering
                "stop_if_fail": False,
                "score_penalty": (overrides or {}).get("score_penalty"),
                "weight": None,
                "parameters": params,
                "terminal_verdict": None,
                "enabled": True,
            }
            junctions.append(j)
    log.info(f"  Injected {len(_CUSHION_PENALTY_STRATEGIES) * len(injections)} "
             f"junctions for cushion penalty bands")


# ── TBD scoring formula capture (OTA-686) ───────────────────────────────
#
# Captures code formulas as definitions for all TBD scoring criteria.
# Canonical formulas are recorded from strategy_scorer.py.
# Proxy formulas are captured AND flagged with their planned replacement.
# We are NOT defining optimal formulas — only capturing what exists today.

_SCORING_FORMULA_DEFINITIONS = {
    # ── Canonical formulas (code has the real formula) ──────────────────
    "credit_width": {
        "condition_expression": "abs(net_debit) / spread_width * 100",
        "intent": (
            "Credit Width %: net credit received as percentage of spread width. "
            "Canonical formula from strategy_scorer.py:222-226. Output: 0–100 (natural %)."
        ),
        "referenced_named_values": ["net_debit", "spread_width"],
        "parameter_schema": None,
    },
    "liquidity": {
        "condition_expression": "long_volume + short_volume + long_oi + short_oi",
        "intent": (
            "Liquidity: sum of both legs' volume and open interest. "
            "Canonical formula from strategy_scorer.py:228-233. "
            "NORMALIZATION OWED: raw sum, not yet on [0,100] scale."
        ),
        "referenced_named_values": ["long_volume", "short_volume", "long_oi", "short_oi"],
        "parameter_schema": None,
    },
    "delta_quality": {
        "condition_expression": "formula:delta_quality",
        "intent": (
            "Delta Quality: gaussian-like peak around target delta range. "
            "Canonical formula from strategy_scorer.py:389-392. "
            "Output: 0–1, multiplied by 100 at scoring. "
            "Parameterised by delta_center and delta_half_range."
        ),
        "referenced_named_values": ["delta"],
        "parameter_schema": {
            "delta_center": {"type": "number", "description": "Peak delta target"},
            "delta_half_range": {"type": "number", "description": "Half-width of gaussian peak"},
        },
    },
    "payout_ratio": {
        "condition_expression": "(delta * underlying_price * 0.10 * 100) / premium_dollars",
        "intent": (
            "Payout Ratio: expected 10% move payout relative to premium paid. "
            "Canonical formula from strategy_scorer.py:399-407. "
            "NORMALIZATION OWED: raw ratio, not yet on [0,100] scale."
        ),
        "referenced_named_values": ["delta", "underlying_price", "premium_dollars"],
        "parameter_schema": None,
    },
    "delta_otm_score": {
        "condition_expression": "1.0 - delta / 0.25",
        "intent": (
            "Delta OTM Score: how far out-of-the-money the option is. "
            "Canonical formula from strategy_scorer.py:394-397. "
            "Output: 0–1 range (0.25 delta → 0, 0 delta → 1). "
            "NORMALIZATION OWED: 0–1 scale, multiply by 100 for [0,100]."
        ),
        "referenced_named_values": ["delta"],
        "parameter_schema": None,
    },
    "bid_ask_tightness": {
        "condition_expression": "1.0 - bid_ask_spread_pct / 100.0",
        "intent": (
            "Bid Ask Tightness: inverse of bid-ask spread percentage. "
            "Canonical formula from strategy_scorer.py:409-412. "
            "Output: 0–1 range. "
            "NORMALIZATION OWED: 0–1 scale, multiply by 100 for [0,100]."
        ),
        "referenced_named_values": ["bid_ask_spread_pct"],
        "parameter_schema": None,
    },

    # ── Proxy formulas (captured as-is, flagged for replacement) ────────
    "theta_gamma_ratio": {
        "condition_expression": "abs(net_theta) / max_loss",
        "intent": (
            "Theta Gamma Ratio: PROXY — currently identical to theta_margin_ratio "
            "(abs(net_theta) / max_loss). NO gamma involved. "
            "PLANNED REPLACEMENT: true theta/gamma ratio requires per-leg gamma "
            "propagated to the candidate by the options-chain adapter (later feature). "
            "NORMALIZATION OWED: raw ratio, not yet on [0,100] scale."
        ),
        "referenced_named_values": ["net_theta", "max_loss"],
        "parameter_schema": None,
    },
    "sma_alignment_score": {
        "condition_expression": "0.5",
        "intent": (
            "SMA Alignment Score: PROXY — 0.5 passthrough (no real formula). "
            "REPLACEMENT: classification(BULLISH/BEARISH/MIXED/NEUTRAL) → [0,1] "
            "via compute_sma_signal(). NORMALIZATION OWED: 0–1 scale × 100."
        ),
        "referenced_named_values": ["sma_8", "sma_21", "sma_50", "sma_alignment_classification"],
        "parameter_schema": None,
    },
    "iv_percentile_cost": {
        "condition_expression": "1.0 - iv_decimal",
        "intent": (
            "IV Percentile Cost: PROXY — linear inversion of raw IV. "
            "PLANNED REPLACEMENT: true IV percentile requires historical-IV producer "
            "(adapter feature, later). Current formula penalises high IV linearly. "
            "NORMALIZATION OWED: 0–1 scale, multiply by 100 for [0,100]."
        ),
        "referenced_named_values": ["iv"],
        "parameter_schema": None,
    },
    "runway_score": {
        "condition_expression": "theta_runway_days",
        "intent": (
            "Runway Score: PROXY — raw theta_runway_days (premium / daily_theta). "
            "PLANNED REPLACEMENT: [0,100] normalization to be defined during tuning "
            "(e.g. sigmoid or min-max over a domain-appropriate range). "
            "NORMALIZATION OWED: raw days, not on [0,100] scale."
        ),
        "referenced_named_values": ["theta_runway_days"],
        "parameter_schema": None,
    },
    "open_interest": {
        "condition_expression": "open_interest",
        "intent": (
            "Open Interest: PROXY — raw open_interest value. "
            "PLANNED REPLACEMENT: [0,100] normalization to be defined during tuning "
            "(e.g. log-scale or percentile rank over recent chain). "
            "NORMALIZATION OWED: raw count, not on [0,100] scale."
        ),
        "referenced_named_values": ["open_interest"],
        "parameter_schema": None,
    },
}


def enrich_scoring_formulas(rules):
    """
    OTA-686: Fill in TBD scoring formula definitions.

    For each scoring criterion with condition_expression == 'TBD', replace with
    the captured formula definition from _SCORING_FORMULA_DEFINITIONS.
    """
    enriched = 0
    for r in rules:
        if r["phase"] != "scoring":
            continue
        ce = r.get("condition_expression") or ""
        if "TBD" not in ce:
            continue
        key = r["rule_key"]
        defn = _SCORING_FORMULA_DEFINITIONS.get(key)
        if defn is None:
            log.warning(f"  No formula definition for TBD scoring rule '{key}' — skipping")
            continue

        r["condition_expression"] = defn["condition_expression"]
        r["intent"] = defn["intent"]
        if defn["referenced_named_values"]:
            r["referenced_named_values"] = defn["referenced_named_values"]
        if defn["parameter_schema"] is not None:
            r["parameter_schema"] = defn["parameter_schema"]
        enriched += 1
        is_proxy = "PROXY" in defn["intent"]
        label = "proxy+flag" if is_proxy else "canonical"
        log.info(f"  Enriched '{key}' ({label})")

    log.info(f"Scoring formula enrichment: {enriched} rules captured")
    return rules


# ── OTA-688: Backfill missing rules + resolve code-only rules ───────────
#
# Three responsibilities:
# 1. Flag dependency-owed rules (no-producer inputs) in intent
# 2. Inject code-only rules not in the workbook (DTE 8-13, prob-asymmetry)
# 3. Record architectural dispositions (narrative validator, WAIT_FOR_EARNINGS,
#    engine-internal weights) as documented decisions — NOT as engine rules.
#
# Items already handled by OTA-683 and NOT re-seeded here:
#   - Chart state confirms direction → OTA-683 decomposed to 2 atomic rules
#   - Earnings Route-4 penalty → OTA-683 decomposed to 4 atomic earnings rules
#   - Cushion graduated penalty → OTA-683 injected adj_cushion_penalty_severe/moderate
#   - Stock Extended in Trade Direction (post-scoring) → OTA-683 decomposed
#
# Architectural dispositions (NOT seeded as rules — recorded here as decisions):
#   - Narrative-grounding validator: app-layer post-Claude quality gate, NOT an
#     engine rule. Lives in evaluation_routes.py:917-1024. The engine evaluates
#     deterministic math; Claude narrative validation is a separate concern.
#   - WAIT_FOR_EARNINGS: NOT a 4th verdict band. It is a halt verdict emitted
#     via terminal_verdict on junction rows for earnings routes 2 & 3 (OTA-711).
#     Route 1 emits PASS (trade dead). Route 4 is non-stopping. The 3 band
#     verdicts (EXECUTE >= 70, WAIT 50-69, PASS < 50) are unaffected.
#     WAIT_FOR_EARNINGS is registered in engine_lookups(screening_verdicts)
#     with kind=HALT_VERDICT so OTA-699 startup validation passes.
#   - VerticalSpreadEngine internal weights (EV35/RR25/Prob20/Liq15/Theta5) and
#     NakedOptionEngine internal weights (Delta30/Theta25/IV20/RR15/Liq10): a
#     legacy second scoring system SUPERSEDED by the strategy junction weights
#     (the engine has one weight set per strategy via engine_strategy_rule_junction).
#     NOT seeded. Actual code removal tracked in later screening-rules feature.

# Rules whose inputs have no producer yet — flag in intent
_DEPENDENCY_FLAGS = {
    # Data completeness gates — input not yet produced at runtime
    "data_completeness_iv_rank": (
        "DEPENDENCY: true IV rank requires a historical-IV producer (adapter feature, "
        "later). Current code uses ATM IV / 0.60 as proxy. Gate records the requirement."
    ),
    "data_completeness_atr_14": (
        "DEPENDENCY: ATR_14 requires a technical-indicator producer (adapter feature, "
        "later). Not computed anywhere in the current pipeline. Gate records the requirement."
    ),
    # Cushion vs ATR depends on ATR_14 producer
    "cushion_vs_atr": (
        "DEPENDENCY: cushion_vs_ATR computation requires ATR_14, which has no producer yet. "
        "Gate records the requirement; will activate when ATR_14 adapter ships."
    ),
    # Spread width tier compliance references the width_configuration lookup (OTA-690)
    "spread_width_tier_compliance": (
        "DEPENDENCY: references width_configuration lookup (seeded from Width Configuration "
        "sheet, OTA-690). Runtime tier resolution not yet built."
    ),
}


def backfill_missing_rules(rules, junctions):
    """
    OTA-688: Backfill missing rules and resolve code-only rules.

    1. Flags dependency-owed rules with no-producer annotations in intent.
    2. Injects code-only rules (DTE 8-13 penalty, probability-asymmetry penalty).
    3. Does NOT re-seed items already handled by OTA-683 (chart state, earnings
       Route-4, cushion penalty, stock extended post-scoring decomposition).
    """
    # ── 1. Flag dependency-owed rules ──────────────────────────────────
    flagged = 0
    for r in rules:
        key = r["rule_key"]
        flag = _DEPENDENCY_FLAGS.get(key)
        if flag:
            existing_intent = r.get("intent") or ""
            if "DEPENDENCY:" not in existing_intent:
                r["intent"] = f"{existing_intent} — {flag}" if existing_intent else flag
                flagged += 1
                log.info(f"  Flagged dependency: {key}")

    # ── 2. Inject code-only rules ─────────────────────────────────────
    # These exist in code but NOT in the workbook. They're injected as
    # synthetic rules with full intent documentation.

    code_only_rules = []
    code_only_junctions = []

    # 2a. DTE 8-13 penalty (-20 points)
    # Source: evaluation_routes.py:634-638
    # All strategies — universal adjustment for near-expiry trades
    dte_penalty_rule = {
        "owner_app_id": "OTA",
        "rule_key": "adj_dte_8_13_penalty",
        "phase": "adjustment",
        "tier": "RAW",
        "intent": (
            "DTE 8-13 penalty: trades with 8-13 DTE receive a -20 point adjustment. "
            "Covers the gap between the 7 DTE hard gate and normal scoring. "
            "Code-only rule captured from evaluation_routes.py:634-638."
        ),
        "condition_expression": "dte >= 8 AND dte <= 13",
        "formula_ref": None,
        "referenced_named_values": ["dte"],
        "parameter_schema": {
            "dte_low": {"type": "number", "default": 8},
            "dte_high": {"type": "number", "default": 13},
            "penalty": {"type": "number", "default": -20},
        },
        "null_semantics": None,
        "enabled": True,
    }
    code_only_rules.append(dte_penalty_rule)

    # DTE penalty applies to all strategies
    for strat_key in ["steady_paycheck", "weekly_grind", "trend_rider", "lottery_ticket"]:
        code_only_junctions.append({
            "rule_key": "adj_dte_8_13_penalty",
            "strategy_key": strat_key,
            "evaluation_order": 0,  # placeholder — OTA-684 sets proper ordering
            "stop_if_fail": False,
            "score_penalty": -20.0,
            "weight": None,
            "parameters": {"dte_low": 8, "dte_high": 13, "penalty": -20},
            "terminal_verdict": None,
            "enabled": True,
        })

    # 2b. Probability-asymmetry penalty (graduated: 0/8/15/25)
    # Source: scoring_factors/asymmetry.py:16-41 (OTA-505)
    # All strategies — graduated penalty based on loss/profit probability ratio
    asym_penalty_rule = {
        "owner_app_id": "OTA",
        "rule_key": "adj_probability_asymmetry_penalty",
        "phase": "adjustment",
        "tier": "COMPUTED",
        "intent": (
            "Probability asymmetry penalty (OTA-505): graduated penalty based on "
            "loss/profit probability ratio. ratio >= 2.0 → -25; >= 1.5 → -15; "
            ">= 1.25 → -8; < 1.25 → 0. Null inputs → 0 (no penalty). "
            "Code-only rule captured from scoring_factors/asymmetry.py:16-41."
        ),
        "condition_expression": "formula:probability_asymmetry_penalty",
        "formula_ref": "formula:probability_asymmetry_penalty",
        "referenced_named_values": ["p_max_loss", "p_max_profit"],
        "parameter_schema": {
            "band_severe": {"type": "number", "default": 2.0, "description": "ratio threshold for -25"},
            "band_high": {"type": "number", "default": 1.5, "description": "ratio threshold for -15"},
            "band_moderate": {"type": "number", "default": 1.25, "description": "ratio threshold for -8"},
            "penalty_severe": {"type": "number", "default": -25},
            "penalty_high": {"type": "number", "default": -15},
            "penalty_moderate": {"type": "number", "default": -8},
        },
        "null_semantics": "FAIL_OPEN",
        "enabled": True,
    }
    code_only_rules.append(asym_penalty_rule)

    # Asymmetry penalty applies to all strategies
    for strat_key in ["steady_paycheck", "weekly_grind", "trend_rider", "lottery_ticket"]:
        code_only_junctions.append({
            "rule_key": "adj_probability_asymmetry_penalty",
            "strategy_key": strat_key,
            "evaluation_order": 0,  # placeholder — OTA-684 sets proper ordering
            "stop_if_fail": False,
            "score_penalty": None,  # graduated — penalty comes from formula evaluation
            "weight": None,
            "parameters": {
                "band_severe": 2.0, "band_high": 1.5, "band_moderate": 1.25,
                "penalty_severe": -25, "penalty_high": -15, "penalty_moderate": -8,
            },
            "terminal_verdict": None,
            "enabled": True,
        })

    # Append to main lists
    rules.extend(code_only_rules)
    junctions.extend(code_only_junctions)

    log.info(f"OTA-688 backfill: {flagged} dependency flags, "
             f"{len(code_only_rules)} code-only rules injected, "
             f"{len(code_only_junctions)} junctions added")
    return rules, junctions


# ── OTA-685: Reconcile sheet-vs-code divergences ────────────────────────

# Each entry: (rule_key, field, note)
# field = "intent" → append to rule intent; "rationale" → append to junction rationale
_DIVERGENCE_NOTES = {
    # --- MISSING in code ---
    "underlying_price_floor": (
        "intent",
        "DIVERGENCE: sheet >= 20 (MISSING in code). No code enforces a stock price floor. "
        "Sheet value seeded; revisit in tuning."
    ),
    # --- Liquidity floors ---
    "per_leg_bid_ask_spread": (
        "intent",
        "DIVERGENCE: sheet 0.15 max; code long_call_engine 0.15 default but strategy_scorer "
        "overrides to 0.50; vertical engine has no bid/ask filter. Sheet value seeded; "
        "relaxed code values noted — revisit in tuning."
    ),
    "per_leg_open_interest_floor": (
        "intent",
        "DIVERGENCE: sheet >= 100; engine defaults 50; strategy_scorer relaxes to 10. "
        "Relaxed for low-volume symbols. Sheet value seeded; revisit in tuning."
    ),
    "per_leg_volume_floor": (
        "intent",
        "DIVERGENCE: sheet >= 500; engine defaults 5; strategy_scorer relaxes to 1. "
        "Largest single divergence in audit. Relaxed for low-volume symbols. Sheet value "
        "seeded; revisit in tuning."
    ),
    # --- DTE boundaries ---
    "minimum_dte": (
        "intent",
        "DIVERGENCE: sheet >= 7; code <= 7 -> PASS (off-by-one at exactly 7 DTE). "
        "Sheet value seeded; boundary note for tuning."
    ),
    "maximum_dte": (
        "intent",
        "DIVERGENCE: sheet <= 60; chain-fetch uses 70; NakedOptionEngine allows 90. "
        "Sheet value seeded; wider code values noted for tuning."
    ),
    # --- Cushion % of price ---
    "cushion_of_price": (
        "intent",
        "DIVERGENCE: sheet defines as hard gate (SP >= 1%, WG >= 1.5%); code implements "
        "as graduated penalty (1%/-20, 2%/-10) with no WG-specific threshold. Sheet gate "
        "values seeded; code penalty captured separately in OTA-688 as layered adjustment "
        "(intentional layering per sheet). Revisit in tuning."
    ),
    # --- Cushion vs ATR ---
    "cushion_vs_atr": (
        "intent",
        "DIVERGENCE: sheet hard gate (SP 1-999, WG 1.5-999); code has TODO comment only. "
        "ATR_14 has no producer yet (options-chain adapter, later feature). Sheet values "
        "seeded; revisit when ATR_14 producer is built."
    ),
    # --- Data completeness gates ---
    "data_completeness_iv_rank": (
        "intent",
        "DIVERGENCE: sheet gates on null IV rank; code substitutes ATM IV proxy "
        "(atm_iv/0.60, default 0.5 when unavailable). True IV rank has NO PRODUCER yet "
        "(options-chain adapter, later feature). Sheet gate seeded; code fallback noted."
    ),
    "data_completeness_delta": (
        "intent",
        "DIVERGENCE: sheet gates on null delta; code substitutes Black-Scholes fallback "
        "for after-hours operation, or treats null as 0. Sheet gate seeded; B-S fallback "
        "noted for tuning."
    ),
    "data_completeness_atr_14": (
        "intent",
        "DIVERGENCE: sheet gates on null ATR_14; ATR_14 has NO PRODUCER yet "
        "(options-chain adapter, later feature). Sheet gate seeded; revisit when producer "
        "is built."
    ),
    # --- Spread width tier compliance ---
    "spread_width_tier_compliance": (
        "intent",
        "DIVERGENCE: sheet has 5-tier width table by price; code uses single hardcoded "
        "max_spread_width = 10. Sheet references OTA-690 width_configuration lookup. "
        "Code's single value noted for tuning."
    ),
    # --- Theta load fraction ---
    "theta_load_fraction": (
        "intent",
        "DIVERGENCE: sheet LT gate theta_total/debit <= 0.5 (MISSING in code). "
        "No code computes theta load fraction. Sheet value seeded; revisit in tuning."
    ),
    # --- Scoring criteria with proxy formulas ---
    "iv_rank": (
        "intent",
        "DIVERGENCE: sheet says IV Rank (RAW); code uses ATM IV proxy (atm_iv/0.60). "
        "True IV rank is percentile-based; code is a ratio. Sheet value seeded; "
        "proxy noted for tuning."
    ),
    "theta_gamma_ratio": (
        "intent",
        "DIVERGENCE: sheet formula TBD; code proxy is identical to theta_margin_ratio "
        "(abs(net_theta)/max_loss — no gamma involved). Sheet value seeded; proxy+flag "
        "enriched by OTA-686."
    ),
    "sma_alignment_score": (
        "intent",
        "DIVERGENCE: sheet formula TBD; code uses client-supplied scalar (default 0.5). "
        "When frontend doesn't send it, all candidates get 0.5 (no-op at 30% weight). "
        "Sheet value seeded; proxy+flag enriched by OTA-686."
    ),
}

# Junction-level rationale notes (strategy-specific divergences)
_JUNCTION_DIVERGENCE_NOTES = {
    # DTE window divergences per strategy
    ("steady_paycheck", "dte_window"): (
        "DIVERGENCE: sheet SP 21-45; code 14-45 (min is 14 in code, 21 in sheet). "
        "Sheet value seeded; revisit in tuning."
    ),
    ("weekly_grind", "dte_window"): (
        "DIVERGENCE: sheet WG 14-20; code 14-21 (max is 21 in code, 20 in sheet). "
        "Sheet value seeded; revisit in tuning."
    ),
    ("trend_rider", "dte_window"): (
        "DIVERGENCE: sheet TR 30-45; code 14-60 (both min and max differ). "
        "Sheet value seeded; revisit in tuning."
    ),
    ("lottery_ticket", "dte_window"): (
        "DIVERGENCE: sheet LT 30-60; code 7-60 (min is 7 in code, 30 in sheet). "
        "Sheet value seeded; revisit in tuning."
    ),
    # Cushion per-strategy divergence
    ("weekly_grind", "cushion_of_price"): (
        "DIVERGENCE: sheet WG >= 1.5%; code uses same 1%/2% graduated penalty as SP "
        "(no WG-specific threshold). Sheet value seeded; revisit in tuning."
    ),
}

# Trade type structure divergences — junction rationale on structure gates
_STRUCTURE_DIVERGENCE_NOTES = {
    ("weekly_grind", "long_call"): (
        "DIVERGENCE: sheet includes LONG_CALL for WG; code/business-rules.md = "
        "credit-only. Sheet value seeded per OTA-687; differs from business-rules.md — "
        "reconcile during tuning."
    ),
    ("weekly_grind", "long_put"): (
        "DIVERGENCE: sheet includes LONG_PUT for WG; code/business-rules.md = "
        "credit-only. Sheet value seeded per OTA-687; differs from business-rules.md — "
        "reconcile during tuning."
    ),
    ("lottery_ticket", "iron_condor"): (
        "DIVERGENCE: sheet includes IRON_CONDOR for LT; code has LONG_CALL + LONG_PUT, "
        "no iron condor. Sheet value seeded per OTA-687; differs from code — "
        "reconcile during tuning."
    ),
    ("lottery_ticket", "long_put"): (
        "DIVERGENCE: sheet LT = LONG_PUT + IRON_CONDOR; code/business-rules.md = "
        "LONG_CALL + LONG_PUT (adds LONG_CALL, omits IRON_CONDOR). Sheet structures "
        "seeded per OTA-687; reconcile during tuning."
    ),
}


def reconcile_divergences(rules, junctions):
    """
    OTA-685: Record sheet-vs-code divergences in rule intent and junction rationale.

    Seeds the sheet value as captured; records each divergence as a one-line note.
    No value-correctness decisions — deferred to tuning via rule-management screen.
    """
    noted_rules = 0
    noted_junctions = 0

    # Apply rule-level divergence notes
    for r in rules:
        key = r["rule_key"]
        if key in _DIVERGENCE_NOTES:
            field, note = _DIVERGENCE_NOTES[key]
            existing = r.get("intent") or ""
            if "DIVERGENCE:" not in existing:
                r["intent"] = f"{existing} | {note}" if existing else note
                noted_rules += 1

    # Apply junction-level divergence notes
    for j in junctions:
        strat = j["strategy_key"]
        rule = j["rule_key"]

        # DTE window / cushion per-strategy notes
        combo = (strat, rule)
        if combo in _JUNCTION_DIVERGENCE_NOTES:
            note = _JUNCTION_DIVERGENCE_NOTES[combo]
            existing = j.get("rationale") or ""
            if "DIVERGENCE:" not in existing:
                j["rationale"] = f"{existing} | {note}" if existing else note
                noted_junctions += 1

        # Structure gate notes
        if combo in _STRUCTURE_DIVERGENCE_NOTES:
            note = _STRUCTURE_DIVERGENCE_NOTES[combo]
            existing = j.get("rationale") or ""
            if "DIVERGENCE:" not in existing:
                j["rationale"] = f"{existing} | {note}" if existing else note
                noted_junctions += 1

    log.info(f"  Divergence notes: {noted_rules} rules, {noted_junctions} junctions annotated")
    return rules, junctions


# ── OTA-684: Gate mechanics (evaluation_order, stop_if_fail, score_penalty) ──

# Gate ordering priority — cheapest/most-decisive first.
# Rules not listed here sort alphabetically after these.
_GATE_ORDER_PRIORITY = [
    # Tier 1: Earnings gates — cheapest hard-kill for SP/WG/TR (RAW/DERIVED, no COMPUTED)
    "earnings_route1_no_viable_window",
    "earnings_route2_wait_post_window",
    "earnings_route3_post_entry_better",
    "earnings_route4_pre_momentum_play",
    # Tier 2: Price floor (RAW)
    "underlying_price_floor",
    # Tier 3: Earnings buffer past expiry (DERIVED)
    "earnings_buffer_past_expiry",
    # Tier 4: Liquidity gates (RAW)
    "per_leg_bid_ask_spread",
    "per_leg_open_interest_floor",
    "per_leg_volume_floor",
    # Tier 5: DTE gates (DERIVED)
    "maximum_dte",
    "minimum_dte",
    "dte_window",
    # Tier 6: Data completeness (RAW — fail fast if data missing)
    "data_completeness_iv_rank",
    "data_completeness_delta",
    "data_completeness_atr_14",
    # Tier 7: Spread structure gates
    "credit_pct_of_width_floor",
    "debit_pct_of_width_ceiling",
    "debit_of_width",
    # Tier 8: Value gates (COMPUTED)
    "total_expected_value",
    # Tier 9: Cushion / width gates
    "cushion_of_price",
    "cushion_vs_atr",
    "spread_width_tier_compliance",
    # Tier 10: Chart state (DERIVED)
    "chart_state_valid_alignment",
    "chart_state_matches_trade_direction",
    # Tier 11: Trade type gates (structure validation)
    "theta_load_fraction",
    "require_credit_spread_structure",
    "require_directional_debit_spread",
    "bull_put_credit",
    "bear_call_credit",
    "bull_call_debit",
    "bear_put_debit",
    "long_call",
    "long_put",
    "iron_condor",
    # Tier 12: Soft gates (non-stopping, recorded with penalty)
    "stock_extended_against_entry",
    "stock_extended_in_trade_direction",
]

# Adjustment ordering priority
_ADJUSTMENT_ORDER_PRIORITY = [
    "adj_dte_8_13_penalty",
    "adj_probability_asymmetry_penalty",
    "adj_cushion_penalty_severe",
    "adj_cushion_penalty_moderate",
    "adj_stock_extended_magnitude",
    "adj_stock_extended_direction_match",
    "adj_sma_alignment_against_trade",
    "adj_mixed_chart_signal_on_directional_strategy",
    "adj_cushion_vs_atr_gte_floor",
    "adj_cushion_vs_atr_lte_ceiling",
    "adj_etf_underlying",
]

# Long-DTE rationale note for earnings junction rows
_LONG_DTE_RATIONALE = (
    "A future 180-360 DTE strategy would set stop_if_fail=false, score_penalty=0 on "
    "this earnings gate — the event passes long before expiry. That strategy needs only "
    "junction rows, no engine code change. This demonstrates the junction-only invariant."
)


# ── OTA-769: Seed per-strategy delta params on delta_quality junctions ────

# Carry-forward values from strategy_scorer.py (exact, not rounded):
#   TR: delta_center = (0.50 + 0.70) / 2 = 0.60, delta_half_range = max(0.10, (0.70 - 0.50) / 2) = 0.10
# These match what the scorer injects at runtime today; seeding them on the
# junction row makes the junction self-contained for OTA-779 (scorer deletion).
_DELTA_QUALITY_PARAMS = {
    "trend_rider": {"delta_center": 0.60, "delta_half_range": 0.10},
}


def populate_delta_quality_params(junctions):
    """
    OTA-769: Set delta_center / delta_half_range on delta_quality junction rows.

    The scoring formula (`scoring_formulas.py:delta_quality`) reads these from
    params with fallbacks of 0.35/0.15 — which match neither TR nor LT.  Today
    the strategy_scorer injects the correct values at runtime, but once OTA-779
    removes the scorer the junction row must carry them.  This seeds the exact
    carry-forward values so behaviour is identical before and after OTA-779.
    """
    patched = 0
    for j in junctions:
        if j["rule_key"] != "delta_quality":
            continue
        strat = j["strategy_key"]
        if strat not in _DELTA_QUALITY_PARAMS:
            continue
        params = j.get("parameters") or {}
        params.update(_DELTA_QUALITY_PARAMS[strat])
        j["parameters"] = params
        patched += 1
        log.info(f"  delta_quality params set for {strat}: {params}")

    log.info(f"  OTA-769: patched {patched} delta_quality junction(s)")
    return junctions


def set_gate_mechanics(rules, junctions):
    """
    OTA-684: Set evaluation_order, stop_if_fail, and score_penalty on all
    gate-phase and adjustment-phase junction rows.

    Guarantees:
    - evaluation_order is unique within each (strategy, phase)
    - Earnings gates fire first (order 1-4) among hard stops for SP/WG/TR
    - Hard Gate rows: stop_if_fail=true, score_penalty=None
    - Soft Gate rows: stop_if_fail=false, score_penalty per sheet
    - Post-Scoring/adjustment rows: stop_if_fail=false, score_penalty per sheet
    - ETF bonus: positive score_penalty (+5, not -5)
    - Long-DTE earnings pattern captured in rationale
    """
    rule_phases = {r["rule_key"]: r["phase"] for r in rules}

    # Build priority index for ordering
    gate_priority = {k: i for i, k in enumerate(_GATE_ORDER_PRIORITY)}
    adj_priority = {k: i for i, k in enumerate(_ADJUSTMENT_ORDER_PRIORITY)}

    def sort_key(j):
        rule_key = j["rule_key"]
        phase = rule_phases.get(rule_key, "")
        if phase == "gate":
            return gate_priority.get(rule_key, 900)
        elif phase == "adjustment":
            return adj_priority.get(rule_key, 900)
        return 900

    # Group junctions by (strategy, phase)
    from collections import defaultdict
    groups = defaultdict(list)
    for j in junctions:
        phase = rule_phases.get(j["rule_key"], "scoring")
        if phase in ("gate", "adjustment"):
            groups[(j["strategy_key"], phase)].append(j)

    # Reassign evaluation_order within each group
    for (strat_key, phase), group in groups.items():
        group.sort(key=sort_key)
        for order, j in enumerate(group, 1):
            j["evaluation_order"] = order

    # Add long-DTE rationale to earnings junction rows
    for j in junctions:
        if j["rule_key"].startswith("earnings_route"):
            existing = j.get("rationale") or ""
            if _LONG_DTE_RATIONALE not in existing:
                j["rationale"] = (
                    f"{existing} | {_LONG_DTE_RATIONALE}" if existing
                    else _LONG_DTE_RATIONALE
                )

    # OTA-711: Clear terminal_verdict on LT earnings routes.
    # LT's long-dated profile means earnings is informational, not a hard halt
    # verdict. LT's stop_if_fail=true with terminal_verdict=NULL is a known gap
    # flagged for follow-up under OTA-680.
    for j in junctions:
        if j["strategy_key"] == "lottery_ticket" and j["rule_key"].startswith("earnings_route"):
            j["terminal_verdict"] = None

    # Verify no duplicates
    order_check = defaultdict(list)
    for j in junctions:
        phase = rule_phases.get(j["rule_key"], "scoring")
        if phase in ("gate", "adjustment"):
            key = (j["strategy_key"], phase, j["evaluation_order"])
            order_check[key].append(j["rule_key"])

    dupes = {k: v for k, v in order_check.items() if len(v) > 1}
    if dupes:
        for k, rules_at_order in dupes.items():
            log.warning(f"  Duplicate evaluation_order {k}: {rules_at_order}")
    else:
        log.info("  No duplicate evaluation_orders detected")

    total_gate = sum(1 for j in junctions if rule_phases.get(j["rule_key"]) == "gate")
    total_adj = sum(1 for j in junctions if rule_phases.get(j["rule_key"]) == "adjustment")
    log.info(f"  Gate mechanics set: {total_gate} gate junctions, {total_adj} adjustment junctions")

    return junctions


# ── OTA-689: Formula registry (scanned from engine_rules.formula_ref) ────

def build_formula_registry(rules: list[dict]) -> list[dict]:
    """Scan all rules for non-null formula_ref values and produce SHARED
    engine_lookups rows under lookup_set='formula_registry'.

    The list is scanned from the rules data (not hand-maintained) so it stays
    correct as OTA-686/OTA-688 add formula names. The engine's startup
    validation (§6.6) checks that every formula_ref in engine_rules has a
    matching entry in this registry.

    OTA-689 re-open: enriched payloads carry intent, inputs, output_type, and
    notes extracted from the rule data. This is one half of the dual-validation
    contract (the other half is the live rule-library registry in Wave-3).
    """
    seen = {}  # formula_name → rule dict (first rule that references it)
    for r in rules:
        ref = r.get("formula_ref")
        if not ref:
            continue
        name = ref.removeprefix("formula:")
        if name not in seen:
            seen[name] = r

    # Per-formula enrichment: intent, inputs, output_type, notes.
    # Extracted from rule data; hand-curated where the rule's intent is sparse.
    _ENRICHMENTS = {
        # ── Gate formulas ──
        "chart_state_matches_direction": {
            "intent": "Chart state alignment must match trade direction (bullish for bull, bearish for bear).",
            "inputs": ["chart_state", "trade_direction"],
            "output_type": "bool",
        },
        "earnings_route1_no_viable_window": {
            "intent": "Earnings Route 1: no viable window — dte_before <= 7 and dte_after < 14. Halt verdict: PASS.",
            "inputs": ["next_earnings_date", "entry_date", "expiry_date", "dte_before_earnings", "dte_after_earnings"],
            "output_type": "bool",
        },
        "earnings_route2_wait_post_window": {
            "intent": "Earnings Route 2: pre-earnings window too short, post-earnings window viable — dte_before <= 7 and dte_after >= 14. Halt verdict: WAIT_FOR_EARNINGS.",
            "inputs": ["next_earnings_date", "entry_date", "expiry_date", "dte_before_earnings", "dte_after_earnings"],
            "output_type": "bool",
        },
        "earnings_route3_post_entry_better": {
            "intent": "Earnings Route 3: post-earnings entry likely better — dte_before >= 8 and dte_after >= 21. Halt verdict: WAIT_FOR_EARNINGS.",
            "inputs": ["next_earnings_date", "entry_date", "expiry_date", "dte_before_earnings", "dte_after_earnings"],
            "output_type": "bool",
        },
        "earnings_route4_pre_momentum_play": {
            "intent": "Earnings Route 4: pre-earnings momentum play — dte_before >= 8 and dte_after < 21. Score with -15 penalty, effective DTE = dte_before - 1.",
            "inputs": ["next_earnings_date", "entry_date", "expiry_date", "dte_before_earnings", "dte_after_earnings"],
            "output_type": "bool",
        },
        # ── Scoring formulas ──
        "bid_ask_tightness": {
            "intent": "Inverse of bid-ask spread percentage. Tighter spreads score higher.",
            "inputs": ["bid_ask_spread_pct"],
            "output_type": "score_0_1",
            "notes": "Normalization owed: 0-1 scale, multiply by 100 for [0,100].",
        },
        "credit_width": {
            "intent": "Net credit received as percentage of spread width.",
            "inputs": ["net_debit", "spread_width"],
            "output_type": "score_0_100",
        },
        "delta_otm_score": {
            "intent": "How far out-of-the-money the option is. 0.25 delta maps to 0; 0 delta maps to 1.",
            "inputs": ["delta"],
            "output_type": "score_0_1",
            "notes": "Normalization owed: 0-1 scale, multiply by 100 for [0,100].",
        },
        "delta_quality": {
            "intent": "Gaussian-like peak around a target delta range. Parameterised by delta_center and delta_half_range.",
            "inputs": ["delta"],
            "output_type": "score_0_1",
            "notes": "Junction parameters: delta_center, delta_half_range.",
        },
        "expected_value": {
            "intent": "Expected value of the trade: (probability of profit * max gain) - (probability of loss * max loss).",
            "inputs": ["p_max_profit", "max_profit", "p_max_loss", "max_loss"],
            "output_type": "decimal",
            "notes": "COMPUTED tier — requires Black-Scholes probability matrix.",
        },
        "iv_percentile_cost": {
            "intent": "Linear inversion of raw IV. Penalises high IV.",
            "inputs": ["iv"],
            "output_type": "score_0_100",
            "notes": "PROXY: true IV percentile requires historical-IV producer (adapter feature, later). Current formula penalises high IV linearly.",
        },
        "iv_rank": {
            "intent": "IV rank as a percentile of historical IV range.",
            "inputs": ["iv_rank"],
            "output_type": "score_0_100",
            "notes": "PROXY: code uses ATM IV / 0.60 as proxy. True IV rank is percentile-based; current implementation is a ratio.",
        },
        "liquidity": {
            "intent": "Combined liquidity from both legs' volume and open interest.",
            "inputs": ["long_volume", "short_volume", "long_oi", "short_oi"],
            "output_type": "decimal",
            "notes": "Normalization owed: raw sum, not yet on [0,100] scale.",
        },
        "open_interest": {
            "intent": "Raw open interest value as a scoring signal.",
            "inputs": ["open_interest"],
            "output_type": "decimal",
            "notes": "PROXY: normalization to [0,100] to be defined during tuning (e.g. log-scale or percentile rank).",
        },
        "payout_ratio": {
            "intent": "Expected 10% move payout relative to premium paid.",
            "inputs": ["delta", "underlying_price", "premium_dollars"],
            "output_type": "decimal",
            "notes": "Normalization owed: raw ratio, not yet on [0,100] scale.",
        },
        "probability_of_profit": {
            "intent": "Probability that the trade expires profitable, derived from option delta.",
            "inputs": ["long_delta", "short_delta"],
            "output_type": "score_0_100",
            "notes": "COMPUTED tier — uses long-leg delta (not 1 - short_delta). See business-rules.md.",
        },
        "reward_risk": {
            "intent": "Ratio of maximum reward to maximum risk for the trade.",
            "inputs": ["max_profit", "max_loss"],
            "output_type": "decimal",
        },
        "runway_score": {
            "intent": "How many days of theta the premium can sustain (premium / daily_theta).",
            "inputs": ["theta_runway_days"],
            "output_type": "decimal",
            "notes": "PROXY: normalization to [0,100] to be defined during tuning (e.g. sigmoid or min-max).",
        },
        "sma_alignment_score": {
            "intent": "Score from SMA alignment classification (BULLISH/BEARISH/MIXED/NEUTRAL).",
            "inputs": ["sma_8", "sma_21", "sma_50", "sma_alignment_classification"],
            "output_type": "score_0_1",
            "notes": "PROXY: 0.5 passthrough. Planned replacement: classification-to-score via compute_sma_signal().",
        },
        "theta_gamma_ratio": {
            "intent": "Ratio of theta decay to gamma risk.",
            "inputs": ["net_theta", "max_loss"],
            "output_type": "decimal",
            "notes": "PROXY: currently identical to theta_margin_ratio (abs(net_theta) / max_loss). True theta/gamma requires per-leg gamma propagation.",
        },
        "theta_margin_ratio": {
            "intent": "Daily theta decay as a fraction of maximum loss (margin at risk).",
            "inputs": ["net_theta", "max_loss"],
            "output_type": "decimal",
        },
        # ── Adjustment formulas ──
        "cushion_penalty_moderate": {
            "intent": "Moderate proximity penalty: cushion >= 1.0% and < 2.0% of underlying price → -10 points.",
            "inputs": ["stock_price", "short_strike"],
            "output_type": "decimal",
        },
        "extension_matches_trade_direction": {
            "intent": "Check if stock extension direction matches trade direction (above SMA for bull, below for bear).",
            "inputs": ["stock_price", "sma_50", "trade_direction"],
            "output_type": "bool",
        },
        "probability_asymmetry_penalty": {
            "intent": "Graduated penalty based on loss/profit probability ratio. ratio >= 2.0 → -25; >= 1.5 → -15; >= 1.25 → -8; < 1.25 → 0.",
            "inputs": ["p_max_loss", "p_max_profit"],
            "output_type": "decimal",
            "notes": "Junction parameters: band_severe (2.0), band_high (1.5), band_moderate (1.25), penalty_severe (-25), penalty_high (-15), penalty_moderate (-8).",
        },
    }

    lookups = []
    for i, name in enumerate(sorted(seen), 1):
        rule = seen[name]
        enrichment = _ENRICHMENTS.get(name, {})

        # Fall back to rule data if no enrichment entry
        intent = enrichment.get("intent", rule.get("intent", ""))
        inputs = enrichment.get("inputs")
        if inputs is None:
            ref_vals = rule.get("referenced_named_values")
            if isinstance(ref_vals, list):
                inputs = [v.get("name", str(v)) if isinstance(v, dict) else str(v) for v in ref_vals]
            else:
                inputs = []
        output_type = enrichment.get("output_type", "decimal")
        notes = enrichment.get("notes")

        payload = {
            "status": "pending",
            "intent": intent,
            "inputs": inputs,
            "output_type": output_type,
            "first_referenced_by": rule["rule_key"],
        }
        if notes:
            payload["notes"] = notes

        lookups.append({
            "owner_app_id": "SHARED",
            "lookup_set": "formula_registry",
            "lookup_key": name,
            "payload": payload,
            "sort_order": i,
        })

    return lookups


# ── Database write ───────────────────────────────────────────────────────

def upsert_rules(cursor, rules: list[dict]) -> dict[str, int]:
    """Upsert rules and return {rule_key: rule_id} mapping."""
    rule_id_map = {}
    for r in rules:
        ref_vals = validate_json(r["referenced_named_values"], f"rule {r['rule_key']} referenced_named_values")
        param_sch = validate_json(r["parameter_schema"], f"rule {r['rule_key']} parameter_schema")

        # Check if exists
        cursor.execute(
            "SELECT rule_id FROM dbo.engine_rules WHERE owner_app_id = ? AND rule_key = ?",
            r["owner_app_id"], r["rule_key"],
        )
        existing = cursor.fetchone()

        if existing:
            rule_id = existing[0]
            cursor.execute("""
                UPDATE dbo.engine_rules SET
                    phase = ?, tier = ?, intent = ?, condition_expression = ?,
                    formula_ref = ?, referenced_named_values = ?, parameter_schema = ?,
                    null_semantics = ?, enabled = ?, updated_at = GETUTCDATE()
                WHERE rule_id = ?
            """, r["phase"], r["tier"], r["intent"], r["condition_expression"],
                r["formula_ref"], ref_vals, param_sch,
                r["null_semantics"], 1 if r["enabled"] else 0, rule_id)
            log.debug(f"  Updated rule: {r['rule_key']} (id={rule_id})")
        else:
            cursor.execute("""
                INSERT INTO dbo.engine_rules
                    (owner_app_id, rule_key, phase, tier, intent, condition_expression,
                     formula_ref, referenced_named_values, parameter_schema, null_semantics, enabled)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, r["owner_app_id"], r["rule_key"], r["phase"], r["tier"], r["intent"],
                r["condition_expression"], r["formula_ref"], ref_vals, param_sch,
                r["null_semantics"], 1 if r["enabled"] else 0)
            cursor.execute("SELECT @@IDENTITY")
            rule_id = int(cursor.fetchone()[0])
            log.debug(f"  Inserted rule: {r['rule_key']} (id={rule_id})")

        rule_id_map[r["rule_key"]] = rule_id

    log.info(f"Rules: {len(rule_id_map)} upserted")
    return rule_id_map


def upsert_strategies(cursor, strategies: list[dict]) -> dict[str, int]:
    """Upsert strategies and return {strategy_key: strategy_id} mapping."""
    strat_id_map = {}
    for s in strategies:
        compat = validate_json(s["compatible_structures"], f"strategy {s['strategy_key']} compatible_structures")
        bands = validate_json(s["verdict_band_set"], f"strategy {s['strategy_key']} verdict_band_set")

        cursor.execute(
            "SELECT strategy_id FROM dbo.engine_strategies WHERE owner_app_id = ? AND strategy_key = ?",
            s["owner_app_id"], s["strategy_key"],
        )
        existing = cursor.fetchone()

        if existing:
            strat_id = existing[0]
            cursor.execute("""
                UPDATE dbo.engine_strategies SET
                    display_name = ?, consumer_surface = ?, description = ?,
                    compatible_structures = ?, verdict_band_set = ?,
                    dte_min = ?, dte_max = ?, enabled = ?, updated_at = GETUTCDATE()
                WHERE strategy_id = ?
            """, s["display_name"], s["consumer_surface"], s["description"],
                compat, bands, s["dte_min"], s["dte_max"],
                1 if s["enabled"] else 0, strat_id)
            log.debug(f"  Updated strategy: {s['strategy_key']} (id={strat_id})")
        else:
            cursor.execute("""
                INSERT INTO dbo.engine_strategies
                    (owner_app_id, strategy_key, display_name, consumer_surface, description,
                     compatible_structures, verdict_band_set, dte_min, dte_max, enabled)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, s["owner_app_id"], s["strategy_key"], s["display_name"],
                s["consumer_surface"], s["description"], compat, bands,
                s["dte_min"], s["dte_max"], 1 if s["enabled"] else 0)
            cursor.execute("SELECT @@IDENTITY")
            strat_id = int(cursor.fetchone()[0])
            log.debug(f"  Inserted strategy: {s['strategy_key']} (id={strat_id})")

        strat_id_map[s["strategy_key"]] = strat_id

    log.info(f"Strategies: {len(strat_id_map)} upserted")
    return strat_id_map


def upsert_junctions(cursor, junctions: list[dict], rule_id_map: dict, strat_id_map: dict):
    """Upsert junction rows keyed on (strategy_id, rule_id)."""
    count = 0
    for j in junctions:
        rule_id = rule_id_map.get(j["rule_key"])
        strat_id = strat_id_map.get(j["strategy_key"])
        if rule_id is None:
            log.warning(f"  Skipping junction: unknown rule_key '{j['rule_key']}'")
            continue
        if strat_id is None:
            log.warning(f"  Skipping junction: unknown strategy_key '{j['strategy_key']}'")
            continue

        params = validate_json(j["parameters"], f"junction {j['strategy_key']}×{j['rule_key']} parameters")

        # terminal_verdict: nullable varchar(32) — validate type + length
        tv = j.get("terminal_verdict")
        if tv is not None:
            tv = str(tv).strip()
            if len(tv) == 0 or len(tv) > 32:
                raise ValueError(
                    f"terminal_verdict must be a non-empty string of length ≤ 32 or null — "
                    f"got {tv!r} (len={len(tv)}) on ({j['strategy_key']}, {j['rule_key']})"
                )
        else:
            tv = None

        cursor.execute(
            "SELECT junction_id FROM dbo.engine_strategy_rule_junction WHERE strategy_id = ? AND rule_id = ?",
            strat_id, rule_id,
        )
        existing = cursor.fetchone()

        if existing:
            jid = existing[0]
            cursor.execute("""
                UPDATE dbo.engine_strategy_rule_junction SET
                    evaluation_order = ?, stop_if_fail = ?, score_penalty = ?,
                    weight = ?, parameters = ?, terminal_verdict = ?,
                    enabled = ?, updated_at = GETUTCDATE()
                WHERE junction_id = ?
            """, j["evaluation_order"], 1 if j["stop_if_fail"] else 0,
                j["score_penalty"], j["weight"], params, tv,
                1 if j["enabled"] else 0, jid)
            log.debug(f"  Updated junction: {j['strategy_key']}×{j['rule_key']}")
        else:
            cursor.execute("""
                INSERT INTO dbo.engine_strategy_rule_junction
                    (strategy_id, rule_id, evaluation_order, stop_if_fail,
                     score_penalty, weight, parameters, terminal_verdict, enabled)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, strat_id, rule_id, j["evaluation_order"],
                1 if j["stop_if_fail"] else 0, j["score_penalty"],
                j["weight"], params, tv, 1 if j["enabled"] else 0)
            log.debug(f"  Inserted junction: {j['strategy_key']}×{j['rule_key']}")
        count += 1

    log.info(f"Junctions: {count} upserted")


def upsert_lookups(cursor, lookups: list[dict]):
    """Upsert lookup rows keyed on (owner_app_id, lookup_set, lookup_key)."""
    count = 0
    for lk in lookups:
        payload = validate_json(lk["payload"], f"lookup {lk['lookup_set']}/{lk['lookup_key']}")

        cursor.execute(
            "SELECT lookup_id FROM dbo.engine_lookups WHERE owner_app_id = ? AND lookup_set = ? AND lookup_key = ?",
            lk["owner_app_id"], lk["lookup_set"], lk["lookup_key"],
        )
        existing = cursor.fetchone()

        if existing:
            lid = existing[0]
            cursor.execute("""
                UPDATE dbo.engine_lookups SET
                    payload = ?, sort_order = ?, enabled = ?
                WHERE lookup_id = ?
            """, payload, lk["sort_order"], 1, lid)
        else:
            cursor.execute("""
                INSERT INTO dbo.engine_lookups
                    (owner_app_id, lookup_set, lookup_key, payload, sort_order, enabled)
                VALUES (?, ?, ?, ?, ?, ?)
            """, lk["owner_app_id"], lk["lookup_set"], lk["lookup_key"],
                payload, lk["sort_order"], 1)
        count += 1

    log.info(f"Lookups: {count} upserted")


# ── OTA-683 follow-up: delete decomposed compounds from DB ───────────────

def cleanup_decomposed_compounds(cursor, compound_rule_keys: set[str]):
    """
    Delete compound rules and their junction rows from the DB.

    The decomposition logic removes compounds from the in-memory seed, but on
    a re-run against an existing DB the old compound rows persist (upsert-only
    model never deletes). This function explicitly removes them.

    Also unbinds `days_until_next_earnings` (delete junctions, set enabled=0)
    because the 4-route earnings tree supersedes it and they share an
    evaluation_order slot.
    """
    # Rules to unbind (delete junctions) but keep as disabled rule rows
    keep_disabled = {"days_until_next_earnings"}

    # Additional compound rules whose atoms already exist but that the decomposer
    # doesn't catch (debit_of_width is standalone, not found by credit/debit search)
    extra_retirements = {"debit_of_width"}
    all_keys = compound_rule_keys | extra_retirements

    if not all_keys:
        log.info("No compound rule_keys to clean up.")
        return

    # Delete junction rows for ALL compounds (including keep_disabled)
    for rk in all_keys:
        cursor.execute("""
            DELETE j FROM dbo.engine_strategy_rule_junction j
            JOIN dbo.engine_rules r ON r.rule_id = j.rule_id
            WHERE r.owner_app_id = 'OTA' AND r.rule_key = ?
        """, rk)
        deleted = cursor.rowcount
        if deleted:
            log.info(f"  Deleted {deleted} junction row(s) for compound '{rk}'")

    # Delete compound rule rows (except keep_disabled — those are upserted as enabled=0)
    for rk in all_keys:
        if rk in keep_disabled:
            continue
        cursor.execute("""
            DELETE FROM dbo.engine_rules
            WHERE owner_app_id = 'OTA' AND rule_key = ?
        """, rk)
        if cursor.rowcount:
            log.info(f"  Deleted rule row: '{rk}'")

    # Unbind days_until_next_earnings: delete any leftover junctions, ensure disabled
    cursor.execute("""
        DELETE j FROM dbo.engine_strategy_rule_junction j
        JOIN dbo.engine_rules r ON r.rule_id = j.rule_id
        WHERE r.owner_app_id = 'OTA' AND r.rule_key = 'days_until_next_earnings'
    """)
    dte_junctions_deleted = cursor.rowcount
    if dte_junctions_deleted:
        log.info(f"  Deleted {dte_junctions_deleted} leftover junction row(s) for 'days_until_next_earnings'")

    cursor.execute("""
        UPDATE dbo.engine_rules SET enabled = 0, updated_at = GETUTCDATE()
        WHERE owner_app_id = 'OTA' AND rule_key = 'days_until_next_earnings'
    """)
    if cursor.rowcount:
        log.info("  Disabled rule 'days_until_next_earnings' (kept for divergence record)")


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed engine_* config tables from Scoring Parameters.xlsx")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Path to workbook")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, do not write to DB")
    args = parser.parse_args()

    if not args.xlsx.exists():
        log.error(f"Workbook not found: {args.xlsx}")
        sys.exit(1)

    log.info(f"Parsing workbook: {args.xlsx}")
    rules, strategies, junctions, lookups = parse_workbook(args.xlsx)

    log.info(f"Parsed: {len(rules)} rules, {len(strategies)} strategies, "
             f"{len(junctions)} junctions, {len(lookups)} lookups")

    # OTA-683: Decompose compound rules into atomic rules
    log.info("Decomposing compound rules into atomic rules (OTA-683)...")
    rules, junctions, compound_removals = decompose_compound_rules(rules, junctions)

    log.info(f"After decomposition: {len(rules)} rules, {len(junctions)} junctions")

    # OTA-686: Fill in TBD scoring formula definitions + proxy flags
    log.info("Enriching TBD scoring formulas (OTA-686)...")
    rules = enrich_scoring_formulas(rules)

    # OTA-688: Backfill missing rules + resolve code-only rules
    log.info("Backfilling missing rules + injecting code-only rules (OTA-688)...")
    rules, junctions = backfill_missing_rules(rules, junctions)

    log.info(f"After backfill: {len(rules)} rules, {len(junctions)} junctions")

    # OTA-685: Record sheet-vs-code divergences in rationale/intent
    log.info("Recording sheet-vs-code divergences (OTA-685)...")
    rules, junctions = reconcile_divergences(rules, junctions)

    # OTA-684: Set gate mechanics (evaluation_order, stop_if_fail, score_penalty)
    log.info("Setting gate mechanics on junction rows (OTA-684)...")
    junctions = set_gate_mechanics(rules, junctions)

    # OTA-769: Seed delta_center / delta_half_range on delta_quality junctions
    log.info("Populating delta_quality params (OTA-769)...")
    junctions = populate_delta_quality_params(junctions)

    # OTA-689: Build formula registry from engine_rules.formula_ref (scanned, not hand-copied)
    log.info("Building formula registry from engine_rules.formula_ref (OTA-689)...")
    formula_lookups = build_formula_registry(rules)
    lookups.extend(formula_lookups)
    log.info(f"Formula registry: {len(formula_lookups)} formulas registered")

    if args.dry_run:
        log.info("DRY RUN — no database writes")
        for r in rules:
            log.info(f"  Rule: {r['rule_key']} ({r['phase']}/{r['tier']})")
        for s in strategies:
            log.info(f"  Strategy: {s['strategy_key']} ({s['consumer_surface']})")
        for j in junctions:
            log.info(f"  Junction: {j['strategy_key']} × {j['rule_key']} "
                     f"(order={j['evaluation_order']}, stop={j['stop_if_fail']})")
        for lk in lookups:
            log.info(f"  Lookup: {lk['owner_app_id']}/{lk['lookup_set']}/{lk['lookup_key']}")
        return

    conn = connect_db()
    cursor = conn.cursor()
    try:
        # Verify engine_apps seed exists
        cursor.execute("SELECT app_id FROM dbo.engine_apps")
        apps = {row[0] for row in cursor.fetchall()}
        if "SHARED" not in apps or "OTA" not in apps:
            log.error("engine_apps missing SHARED or OTA rows. Apply OTA-681 migration first.")
            sys.exit(1)

        rule_id_map = upsert_rules(cursor, rules)
        strat_id_map = upsert_strategies(cursor, strategies)
        upsert_junctions(cursor, junctions, rule_id_map, strat_id_map)
        upsert_lookups(cursor, lookups)

        # OTA-683 follow-up: remove decomposed compounds + unbind days_until_next_earnings
        log.info("Cleaning up decomposed compound rules (OTA-683 follow-up)...")
        cleanup_decomposed_compounds(cursor, compound_removals)

        conn.commit()
        log.info("Seed import committed successfully.")

        # Summary counts
        for table in ["engine_rules", "engine_strategies", "engine_strategy_rule_junction", "engine_lookups"]:
            cursor.execute(f"SELECT COUNT(*) FROM dbo.{table}")
            log.info(f"  {table}: {cursor.fetchone()[0]} rows")

    except Exception:
        conn.rollback()
        log.exception("Seed import FAILED — rolled back")
        sys.exit(1)
    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    main()
